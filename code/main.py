from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List
from openpyxl import load_workbook, Workbook
import os

origins = [
    "http://127.0.0.1:5500",  
]
app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,  
    allow_credentials=True,
    allow_methods=["*"],  
    allow_headers=["*"],  
)

# Excel file path (you can place it in the root or a specific folder)
EXCEL_FILE_PATH = "../data/SE347.P12.xlsx"

# Pydantic Model for class
class ClassItem(BaseModel):
    id: int
    name: str

class NewClassItem(BaseModel):
    name: str


def load_students_from_excel(index: int, limit: int = -1):
    if not os.path.exists(EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Name"])  
        wb.save(EXCEL_FILE_PATH)
        return []

    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb.active

    if limit == -1:
        limit = ws.max_row - 1
    class_data = []
    start_row = index + 2  
    end_row = start_row + limit

    for row in ws.iter_rows(min_row=start_row, max_row=end_row - 1, values_only=True):
        if row[0] is not None:  
            class_data.append({"id": row[0], "name": row[1]})
    return class_data


# Helper function to save data to Excel file
def save_students_to_excel(class_data):
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Name"])  # Add headers

    for cls in class_data:
        ws.append([cls["id"], cls["name"]])

    wb.save(EXCEL_FILE_PATH)


@app.get("/students/", response_model=List[ClassItem])
async def get_students(index: int = 0, limit: int = 5):
    class_data = load_students_from_excel(index, limit)
    return class_data


# Delete a class
@app.delete("/students/{student_id}")
async def delete_class(student_id: int):
    class_data = load_students_from_excel(0)

    # Filter out the class to be deleted
    new_class_data = [cls for cls in class_data if cls["id"] != student_id]

    if len(new_class_data) == len(class_data):
        raise HTTPException(status_code=404, detail="Class not found")

    # # Save updated data back to Excel
    save_students_to_excel(new_class_data)

    return {"message": "Student deleted successfully"}


# Edit a class
@app.put("/students/{class_id}", response_model=ClassItem)
async def edit_class(class_id: int, updated_class: NewClassItem):
    class_data = load_students_from_excel(0)

    for cls in class_data:
        if cls["id"] == class_id:
            cls["name"] = updated_class.name
            # Save updated data back to Excel
            save_students_to_excel(class_data)
            return cls

    raise HTTPException(status_code=404, detail="Class not found")
